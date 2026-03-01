import os
from io import BytesIO
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generate_order_excel(order):
    """Generate Excel file for order"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Details"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="C48F1E", end_color="C48F1E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=16, color="C48F1E")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "AL BURAQ GROUP - Order Details"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Order Info
    ws['A3'] = "Order Number:"
    ws['B3'] = order.order_number
    ws['A4'] = "Date:"
    ws['B4'] = order.created_at.strftime("%Y-%m-%d %H:%M")
    ws['A5'] = "Customer:"
    ws['B5'] = order.customer_name
    ws['A6'] = "Phone:"
    ws['B6'] = order.customer_phone
    ws['A7'] = "Email:"
    ws['B7'] = order.customer_email or "N/A"
    ws['A8'] = "Company:"
    ws['B8'] = order.customer_company or "N/A"
    ws['A9'] = "Country:"
    ws['B9'] = order.customer_country or "N/A"
    
    # Bold labels
    for row in range(3, 10):
        ws[f'A{row}'].font = Font(bold=True)
    
    # Headers for items
    headers = ['#', 'Category', 'Product Name', 'Quantity', 'Unit', 'Price (RMB)', 'Total (RMB)']
    header_row = 11
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Items
    current_row = header_row + 1
    for idx, item in enumerate(order.items.all(), 1):
        data = [
            idx,
            item.category_name,
            item.product_name,
            item.quantity,
            item.unit,
            float(item.unit_price),
            float(item.get_total())
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            if col in [6, 7]:  # Price columns
                cell.number_format = '¥#,##0.00'
        
        current_row += 1
    
    # Summary
    summary_row = current_row + 1
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws[f'A{summary_row}'] = "SUBTOTAL:"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'A{summary_row}'].alignment = Alignment(horizontal="right")
    ws[f'F{summary_row}'] = float(order.subtotal)
    ws[f'F{summary_row}'].font = Font(bold=True)
    ws[f'F{summary_row}'].number_format = '¥#,##0.00'
    
    # Total items
    ws[f'A{summary_row + 1}'] = f"Total Items: {order.total_items}"
    ws[f'A{summary_row + 1}'].font = Font(bold=True)
    
    # Notes
    if order.notes:
        ws[f'A{summary_row + 3}'] = "Notes:"
        ws[f'A{summary_row + 3}'].font = Font(bold=True)
        ws.merge_cells(f'A{summary_row + 4}:F{summary_row + 4}')
        ws[f'A{summary_row + 4}'] = order.notes
    
    # Adjust column widths
    column_widths = [5, 20, 40, 10, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create Django file
    filename = f"order_{order.order_number}.xlsx"
    return ContentFile(output.read(), name=filename)
